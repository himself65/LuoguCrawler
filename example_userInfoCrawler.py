#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
注意：
    start_num end_num为需手动填写 
"""
from luogu import *
from openpyxl import Workbook
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from urllib import request, error
from queue import Queue

import time
import queue
import os
import ssl
import json
import threading
import linecache

# 必填内容
start_num = 1
end_num = 1000

# 洛谷网站
defaultURL = "https://www.luogu.org"
userURL = "https://www.luogu.org/space/show?uid="
# 此处不建议修改
title = ['id', '名字', '头像', '总提交数', 'AC数', '贡献', '活跃', '积分', '用户类型', '注册时间']
wbName = 'luogu2.xlsx'
wsName = '1'
downloadPath = 'download/'
imagePath = downloadPath + 'img/'
taskPath = downloadPath + 'task/'


def download_img(url, userName):
    """ 下载图片到download/文件夹下
    """
    loc = imagePath + userName + '.png'
    if os.path.exists(loc):
        return
    try:
        # 下载图片
        request.urlretrieve(url, filename=loc)
    except:
        print("\n无法下载文件")


def crawler(taskque, que):
    """ get task from taskque
    """
    try:
        # Init browser
        browser = LuoguBrowser()
        browser.openURL(defaultURL)
    except Exception as e:
        print("无法创建")
        print(e)
        return
    while True:
        try:
            i = taskque.get(block=True, timeout=1)
        except queue.Empty:
            print('无更多任务')
            print('请等待结束')
            return
        try:
            # Get messageURL
            messageURL = userURL + str(i)
            ## View Web
            browser.openURL(messageURL)
            ## getData
            html = browser.getData()
            html = LuoguBrowser.ungzip(html)
            soup = BeautifulSoup(html, 'html.parser')
            # print(soup)
            board = soup.find(
                'ul', {'class': 'am-list am-list-static lg-summary-list'})

            items = board.find_all("li")
            # 0
            userName = soup.find('span', {'name': 'username'}).get_text()
            avatar = items[0].find('img')['src']
            # 1
            allPost = items[1].find_all('span', {'class': 'lg-bignum-num'})
            Num = allPost[0].get_text()
            ACNum = allPost[1].get_text()
            # 2
            Acts = items[4].find('span', {'class': 'lg-right'}).get_text()
            Acts = Acts.split('/')
            contribute = Acts[0]
            active = Acts[1]
            integral = Acts[2]
            # 3
            Type = items[5].find('span', {'class': 'lg-right'}).get_text()
            # 4
            registeredTime = items[6].find('span', {
                'class': 'lg-right'
            }).get_text()
            # make t
            t = [
                i, userName, avatar, Num, ACNum, contribute, active, integral,
                Type, registeredTime
            ]
            # 下载图片
            download_img(avatar, str(i))
            # finish
            taskque.task_done()
            que.put(t)
        except AttributeError:
            que.put([i, '无此人'])
            print('找不到id:', i)
        except Exception as e:
            print(e)


def saveThread(que, sheet):
    while True:
        try:
            t = que.get(block=True, timeout=60)
            if t[1] != '-1':
                sheet.append(t)
                path = taskPath + str(t[0])
                if os.path.exists(path):
                    os.remove(path)
        except queue.Empty:
            return
        que.task_done()


def getLine(num):
    """ 返回是否为true
    """
    if os.path.exists(taskPath + str(num)):
        return True
    return False


def getTaskThread(que, filePath):
    """ 创建任务列队
    """
    # thread = threading.current_thread()
    tgroup = os.listdir(taskPath)
    for item in tgroup:
        try:
            que.put(int(item))
        except ValueError:
            print(item)
    print('剩余任务数量:', que.qsize())


def init():
    print('初始化中')
    if not os.path.exists(downloadPath):
        print('正在创建文件夹download...')
        os.makedirs(downloadPath)
        print('done...')
    if not os.path.exists(taskPath):
        print('正在创建task文件')
        os.makedirs(taskPath)
        # 第一次跑脚本时候使用
        taskMaker(start=start_num, end=end_num)
        print('done...')
    if not os.path.exists(imagePath):
        print('正在创建文件夹image...')
        os.makedirs(imagePath)
        print('done...')
    if not os.path.exists(wbName):
        print('正在创建Excel...')
        wb = Workbook()
        wb.create_sheet(title=wsName)
        wb.save(wbName)
        print('done...')
    print('初始化完成')


def taskMaker(start=1, end=100):
    """ 初始化任务表
    """
    if not os.path.exists(taskPath):
        os.makedirs(taskPath)
    for i in range(start, end):
        f = open(taskPath + str(i), mode='w')
        f.close()
    return


def backgroundThread(saveQue, taskQue):
    while True:
        sz = saveQue.qsize()
        print('待保存量:', sz)
        sz = taskQue.qsize()
        print('剩余任务:', sz)
        time.sleep(30)


def main():
    # MARK -- 参考答案：https://stackoverflow.com/questions/27835619/urllib-and-ssl-certificate-verify-failed-error
    ssl._create_default_https_context = ssl._create_unverified_context
    # init
    init()
    # load data
    print('loading...')
    wb = load_workbook(wbName)
    sheet = wb[wsName]
    sheet.append(title)

    # thread
    saveQue = Queue()
    taskQue = Queue()
    thread = []
    for i in range(0, 9):  # 爬虫线程列队
        t = threading.Thread(
            target=crawler, name=str(i), args=(taskQue, saveQue))
        thread.append(t)
    st = threading.Thread(
        target=saveThread, name='saveThread', args=(saveQue, sheet))
    gt = threading.Thread(
        target=getTaskThread, name='getTaskThread', args=(taskQue, taskPath))
    bg = threading.Thread(
        target=backgroundThread,
        name='backgroundThread',
        args=(saveQue, taskQue))
    print('loading...')
    try:
        print('start!')
        gt.start()
        gt.join()
        for t in thread:
            t.start()
        st.start()
        bg.start()
        st.join()
    except:
        print("线程错误")
    finally:
        wb.save(wbName)


if __name__:
    main()